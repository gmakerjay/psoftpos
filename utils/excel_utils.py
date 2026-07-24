# -*- coding: utf-8 -*-
"""
ฟังก์ชันสำหรับจัดการ Excel (Import/Export)
"""


from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from pathlib import Path


class ExcelManager:
    """จัดการการ Import/Export ข้อมูล Excel สำหรับระบบ POS"""
    
    @staticmethod
    def get_display_width(val):
        """คำนวณความกว้างที่เหมาะสมสำหรับอักขระไทยและ Unicode ใน Excel"""
        if val is None:
            return 0
        s = str(val)
        width = 0.0
        for char in s:
            if ord(char) > 127:
                width += 1.6  # น้ำหนักอักขระไทย / Wide Unicode
            else:
                width += 1.05
        return width

    @staticmethod
    def export_to_excel(data, columns, filename, sheet_name="Sheet1", title=None):
        """
        Export ข้อมูลเป็นไฟล์ Excel โดยปรับความกว้างคอลัมน์ให้อัตโนมัติ ฟอนต์ตัวหนังสือไม่ล้นนอกสเกลช่อง 100%
        """
        try:
            from openpyxl.utils import get_column_letter
            
            # สร้าง workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # สไตล์ส่วนต่างๆ
            title_font = Font(name='Sarabun', size=16, bold=True, color="1E3A8A")
            date_font = Font(name='Sarabun', size=10, italic=True, color="64748B")
            
            header_font = Font(name='Sarabun', size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            
            data_font = Font(name='Sarabun', size=11, color="0F172A")
            even_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )
            
            align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
            align_right = Alignment(horizontal='right', vertical='center')
            
            current_row = 1
            num_cols = len(columns)
            max_col_letter = get_column_letter(num_cols)
            
            # เพิ่มหัวข้อรายงาน
            if title:
                ws.merge_cells(f'A1:{max_col_letter}1')
                title_cell = ws['A1']
                title_cell.value = title
                title_cell.font = title_font
                title_cell.alignment = align_center
                ws.row_dimensions[1].height = 32
                current_row = 2
                
                # วันที่สร้างรายงาน
                ws.merge_cells(f'A2:{max_col_letter}2')
                date_cell = ws['A2']
                date_cell.value = f"วันที่ออกรายงาน: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
                date_cell.font = date_font
                date_cell.alignment = align_center
                ws.row_dimensions[2].height = 20
                current_row = 4
            
            # เพิ่ม Header
            ws.row_dimensions[current_row].height = 28
            for col_idx, column in enumerate(columns, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = str(column)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
                cell.border = thin_border
            
            # เพิ่ม ข้อมูล (Data rows)
            data_start_row = current_row + 1
            for row_offset, row_data in enumerate(data):
                r_idx = data_start_row + row_offset
                ws.row_dimensions[r_idx].height = 24
                row_fill = even_row_fill if row_offset % 2 == 0 else odd_row_fill
                
                for col_idx, column_key in enumerate(columns, start=1):
                    cell = ws.cell(row=r_idx, column=col_idx)
                    
                    if isinstance(row_data, dict):
                        val = row_data.get(column_key, "")
                    else:
                        val = row_data[col_idx - 1] if col_idx <= len(row_data) else ""
                    
                    cell.value = val
                    cell.font = data_font
                    cell.fill = row_fill
                    cell.border = thin_border
                    
                    # Formatting & Alignment ตามประเภทข้อมูล
                    if isinstance(val, (int, float)):
                        cell.alignment = align_right
                        if isinstance(val, float) or (isinstance(val, int) and any(kw in str(column_key).lower() for kw in ['ราคา', 'ยอด', 'ส่วนลด', 'ภาษี', 'บาท', 'total', 'subtotal', 'price', 'discount', 'amount'])):
                            cell.number_format = '#,##0.00'
                        else:
                            cell.number_format = '#,##0'
                    else:
                        str_val = str(val).strip()
                        # จัดกึ่งกลางสำหรับ วันที่/เวลา, รหัส, สถานะ, บาร์โค้ด, เบอร์โทร
                        if any(kw in str(column_key) for kw in ['วันที่', 'เวลา', 'เลขที่', 'รหัส', 'บาร์โค้ด', 'สถานะ', 'หน่วย', 'วิธีชำระ', 'date', 'status']):
                            cell.alignment = align_center
                        else:
                            cell.alignment = align_left
            
            # ปรับความกว้างคอลัมน์ให้อัตโนมัติ (Auto-fit Width with Thai Unicode weighting)
            for col_idx in range(1, num_cols + 1):
                col_letter = get_column_letter(col_idx)
                max_width = 0.0
                
                for cell in ws[col_letter]:
                    w = ExcelManager.get_display_width(cell.value)
                    if w > max_width:
                        max_width = w
                
                # เพิ่ม Safety padding 6 ช่อง และกำหนดความกว้างขั้นต่ำ 16 สูงสุด 70
                final_width = max(max_width + 6.0, 16.0)
                final_width = min(final_width, 70.0)
                ws.column_dimensions[col_letter].width = final_width
            
            # สร้างโฟลเดอร์สำหรับบันทึกหากยังไม่มี
            file_path = Path(filename)
            file_path.parent.mkdir(parents=True, exist_ok=True)
            
            wb.save(str(file_path))
            return True
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False
    
    @staticmethod
    def import_from_excel(filename, sheet_name=0, header_row=0):
        """
        Import ข้อมูลจากไฟล์ Excel
        
        Args:
            filename: ชื่อไฟล์
            sheet_name: ชื่อหรือ index ของ sheet
            header_row: แถวที่เป็น header (0-based)
            
        Returns:
            list of dict
        """
        try:
            import pandas as pd
            clean_path = str(Path(filename).resolve())
            df = pd.read_excel(clean_path, sheet_name=sheet_name, header=header_row)
            return df.to_dict('records')
        except Exception as e:
            print(f"Error importing from Excel: {e}")
            raise e
    
    @staticmethod
    def export_products_template():
        """สร้าง template สำหรับ import สินค้า"""
        columns = [
            "บาร์โค้ด",
            "ชื่อสินค้า",
            "หมวดหมู่",
            "ราคาทุน",
            "ราคาขายปกติ",
            "ราคาขายส่ง",
            "ราคาพิเศษ1",
            "ราคาพิเศษ2",
            "จำนวนสต็อก",
            "สต็อกขั้นต่ำ",
            "หน่วย"
        ]
        
        # ข้อมูลตัวอย่าง
        sample_data = [
            {
                "บาร์โค้ด": "8851234567890",
                "ชื่อสินค้า": "สินค้าตัวอย่าง 1",
                "หมวดหมู่": "อาหารและเครื่องดื่ม",
                "ราคาทุน": 50.00,
                "ราคาขายปกติ": 80.00,
                "ราคาขายส่ง": 70.00,
                "ราคาพิเศษ1": 75.00,
                "ราคาพิเศษ2": 65.00,
                "จำนวนสต็อก": 100,
                "สต็อกขั้นต่ำ": 10,
                "หน่วย": "ชิ้น"
            }
        ]
        
        filename = f"template_สินค้า_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return ExcelManager.export_to_excel(
            sample_data,
            columns,
            filename,
            "สินค้า",
            "Template สำหรับ Import สินค้า"
        )


def export_sales_report(sales_data, filename=None):
    """
    Export รายงานยอดขาย
    
    Args:
        sales_data: ข้อมูลการขาย (list of dict)
        filename: ชื่อไฟล์ (ถ้าไม่ระบุจะสร้างอัตโนมัติ)
    """
    if filename is None:
        filename = f"รายงานยอดขาย_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    columns = [
        "เลขที่ใบเสร็จ",
        "วันที่",
        "ลูกค้า",
        "ยอดรวม",
        "ส่วนลด",
        "ยอดสุทธิ",
        "พนักงานขาย",
        "สถานะ"
    ]
    
    manager = ExcelManager()
    return manager.export_to_excel(
        sales_data,
        columns,
        filename,
        "รายงานยอดขาย",
        "รายงานยอดขาย"
    )


def export_products_list(products_data, filename=None):
    """Export รายการสินค้า"""
    if filename is None:
        filename = f"รายการสินค้า_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    columns = [
        "บาร์โค้ด",
        "ชื่อสินค้า",
        "หมวดหมู่",
        "ราคาทุน",
        "ราคาขาย",
        "สต็อก",
        "สถานะ"
    ]
    
    manager = ExcelManager()
    return manager.export_to_excel(
        products_data,
        columns,
        filename,
        "รายการสินค้า",
        "รายการสินค้าทั้งหมด"
    )


if __name__ == "__main__":
    # ทดสอบ
    print("สร้าง Template สำหรับ Import สินค้า...")
    if ExcelManager.export_products_template():
        print("สร้าง Template สำเร็จ!")
    else:
        print("สร้าง Template ไม่สำเร็จ")
